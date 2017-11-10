from openpyxl import Workbook
from openpyxl.compat import range
#from openpyxl.cell import get_column_letter

## 其实这个demo源于官方文档
## http://openpyxl.readthedocs.io/en/default/usage.html?highlight=get_column_letter
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter   ### openpyxl新版本中使用utils
wb = Workbook()
dest_filename = 'empty_book4.xlsx'
ws1 = wb.active  # 第一个表
ws1.title = "range names"  # 第一个表命名
# 遍历第一个表的1到40行，赋值一个600内的随机数
for row in range(1, 40):
    ws1.append(range(60))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):    ### 10 ~20行
    for col in range(27, 54):  ###  27~54列  即AA~BA列
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))   ### 在当前单元格写入数据，写入的数据就是列号 ，仔细看创建出来的表
print(ws3['AA10'].value)  ###renbin.guo added 
wb.save(filename=dest_filename)
