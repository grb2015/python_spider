'''
	china_offical_carrefour_format.xlsx统计了每个地级市家乐福的数数目，其它类推。
	files = ['china_offical_carrefour_format.xlsx','china_offical_markets_walmat_format.xlsx','china_offical_markets_rt_format.xlsx'\
		,'china_offical_metro_format.xlsx','china_offical_yh_format.xlsx' ,'china_offical_huarun_format.xlsx']

	现在的目的是统计每个地级市各种品牌超市之和。
'''

from openpyxl import load_workbook
import codecs,csv

def cal_total(files):
	for file in files:
		print('-----------file is: -----------------------------------------',file)
		wb = load_workbook(file)
		sheet = wb.get_sheet_by_name('Sheet0')  

		global dict_citys
		# 因为按行，所以返回A1, B1, C1这样的顺序
		i = 0
		for row in sheet.rows:  ### 第一行不要
			if i == 0:
				i = i+1
				continue
			print(row[0].value) 
			key = row[0].value   ### 城市名
			value = row[1].value ### 超市数量

			if key  in dict_citys:  ### 如果这个key存在了
				dict_citys[key] = dict_citys[key] + value
			else:
				dict_citys[key] =  value

			#for cell in row:
			#	print(cell.value)
		print('-----------over file is: -----------------------------------------',file)
### 从超市的excel文件中获取city城市所有的超市总数，比如家乐福数，沃尔玛数，返回一个list,顺序为files的顺序
def get_all_market_count_of_this_city(city,files):
	counts  = []
	for file in files:
		print('-----------file = : -----------------------------------------',file)
		wb = load_workbook(file)
		sheet = wb.get_sheet_by_name('Sheet0') 
		flag = 0
		for row in sheet.rows:
			if row[0].value == city:
				flag = 1
				print('###flag == 1,count = ',row[1].value)

				counts.append(row[1].value)
				break;
		if flag == 0:
			print('#### flag = 0 ')
			counts.append(0)  ### 如果当前file中没有当前城市，则该城市的该超市设置为0
		print('-----------over file = : -----------------------------------------',file)
	print('################ counts = ',counts)
	return counts
if __name__ == '__main__':
	market_file =  codecs.open('china_offical_markets_total.csv', 'w+', encoding='utf-8') 
	writer = csv.writer(market_file)
	writer.writerow(["城市","家乐福数","沃尔玛数","大润发数","麦德龙数","永辉数","华润万家数"])
	#writer.writerow(["城市","超市总数"])

	files = ['china_offical_carrefour_format.xlsx','china_offical_markets_walmat_format.xlsx','china_offical_markets_rt_format.xlsx'\
		,'china_offical_metro_format.xlsx','china_offical_yh_format.xlsx' ,'china_offical_huarun_format.xlsx']
	dict_citys = {}
	cal_total(files)
		#citys = citys[1:]
	print(dict_citys)
	for  key in dict_citys:     ### key代表没一个城市
		info_list = []
		info_list.append(key)
		info_list.append(dict_citys[key])	
		print('current city is :',key)
		
		markets_count = []
		markets_count = get_all_market_count_of_this_city(key,files)
		for count in markets_count:
			info_list.append(count)
		print('### info_list = ',info_list)
		writer.writerow(info_list)

	market_file.close()
